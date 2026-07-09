"""
Data Manager -- Deduplication, merging, and export to CSV / JSON.
"""

import os
import csv
import json
import logging
import re
from datetime import datetime

from config import OUTPUT_DIR, OUTPUT_CSV, OUTPUT_JSON, OUTPUT_XLSX

logger = logging.getLogger(__name__)

# Column order for CSV
CSV_COLUMNS = [
    "name",
    "classification",
    "rating",
    "review_count",
    "category",
    "address_short",
    "address_full",
    "phone",
    "website",
    "hours",
    "price_level",
    "plus_code",
    "latitude",
    "longitude",
    "google_maps_url",
    "photos_count",
    "search_query",
    "area",
    "scraped_at",
    "image_url",
    "review_tags",
]


def classify_listing(item: dict) -> str:
    """
    Classify a listing into:
    - 'women'
    - 'men'
    - 'colive'
    - 'both' (Hostel/PG both)
    - 'remaining' (unclassified / other)
    """
    name = item.get("name", "").strip().lower()
    cat = item.get("category", "").strip().lower()
    addr_full = item.get("address_full", "").strip().lower()
    addr_short = item.get("address_short", "").strip().lower()
    query = item.get("search_query", "").strip().lower()

    # Combined search context
    context = f"{name} {cat} {addr_full} {addr_short} {query}"

    # Check for Co-living first
    colive_keywords = ["colive", "co-live", "coliving", "co-living", "unisex", "mix", "mixed", "couple", "co live"]
    if any(kw in context for kw in colive_keywords):
        return "colive"

    # Check for Women
    women_keywords = ["women", "girl", "ladies", "womens", "girls", "female", "lady", "she", "women's", "girl's"]
    if any(re.search(rf"\b{kw}\b", context) for kw in women_keywords):
        return "women"

    # Check for Men
    men_keywords = ["men", "boy", "gents", "mens", "boys", "male", "he", "men's", "boy's", "gentlemen"]
    if any(re.search(rf"\b{kw}\b", context) for kw in men_keywords):
        return "men"

    # Check for Hostel/PG both
    is_pg = "pg" in context or "paying guest" in context
    is_hostel = "hostel" in context or "hostels" in context

    if (is_pg and is_hostel) or ("both" in context):
        return "both"
    elif is_pg or is_hostel:
        return "both"

    return "remaining"


def deduplicate(results: list) -> list:
    """
    Remove duplicate listings based on normalised name + address.
    Keeps the first (most complete) occurrence.
    """
    seen = set()
    unique = []

    for item in results:
        # Build a dedup key from name + address (normalised)
        name = item.get("name", "").strip().lower()
        addr = item.get("address_full", "").strip().lower()

        # Fallback to short address if full is empty
        if not addr:
            addr = item.get("address_short", "").strip().lower()

        key = f"{name}||{addr}"

        if key not in seen and name:  # skip empty names
            seen.add(key)
            # Classify
            item["classification"] = classify_listing(item)
            unique.append(item)

    removed = len(results) - len(unique)
    if removed > 0:
        logger.info(f"[CLEAN] Deduplication: removed {removed} duplicates, {len(unique)} unique listings remain")
    return unique


def merge_results(existing: list, new_results: list) -> list:
    """
    Merge new results into existing results, deduplicating in the process.
    New results with more data take priority over existing ones.
    """
    combined = list(existing)

    for new_item in new_results:
        new_name = new_item.get("name", "").strip().lower()
        found = False

        for i, existing_item in enumerate(combined):
            if existing_item.get("name", "").strip().lower() == new_name:
                # Merge: fill in blanks from new data
                for key in CSV_COLUMNS:
                    if not existing_item.get(key) and new_item.get(key):
                        combined[i][key] = new_item[key]
                found = True
                break

        if not found and new_name:
            combined.append(new_item)

    return deduplicate(combined)


def export_csv(results: list, output_dir: str = None) -> str:
    """
    Export results to CSV file. Also exports segregated CSV files by classification.
    """
    out_dir = output_dir or OUTPUT_DIR
    os.makedirs(out_dir, exist_ok=True)

    # Master export
    filepath = os.path.join(out_dir, OUTPUT_CSV)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in results:
            writer.writerow(row)
    logger.info(f"[CSV] CSV exported: {filepath} ({len(results)} rows)")

    # Segregated exports
    categories = ["women", "men", "colive", "both", "remaining"]
    for cat in categories:
        cat_results = [r for r in results if r.get("classification") == cat]
        cat_filename = OUTPUT_CSV.replace(".csv", f"_{cat}.csv")
        cat_filepath = os.path.join(out_dir, cat_filename)
        with open(cat_filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            for row in cat_results:
                writer.writerow(row)
        logger.debug(f"[CSV] Segregated CSV exported for '{cat}': {cat_filepath} ({len(cat_results)} rows)")

    return filepath


def export_json(results: list, output_dir: str = None) -> str:
    """
    Export results to JSON file. Also exports segregated JSON files by classification.
    """
    out_dir = output_dir or OUTPUT_DIR
    os.makedirs(out_dir, exist_ok=True)

    filepath = os.path.join(out_dir, OUTPUT_JSON)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(
            {
                "metadata": {
                    "total_listings": len(results),
                    "exported_at": datetime.now().isoformat(),
                    "source": "Google Maps",
                },
                "listings": results,
            },
            f,
            indent=2,
            ensure_ascii=False,
        )
    logger.info(f"[JSON] JSON exported: {filepath} ({len(results)} listings)")

    # Segregated exports
    categories = ["women", "men", "colive", "both", "remaining"]
    for cat in categories:
        cat_results = [r for r in results if r.get("classification") == cat]
        cat_filename = OUTPUT_JSON.replace(".json", f"_{cat}.json")
        cat_filepath = os.path.join(out_dir, cat_filename)
        with open(cat_filepath, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "metadata": {
                        "category": cat,
                        "total_listings": len(cat_results),
                        "exported_at": datetime.now().isoformat(),
                    },
                    "listings": cat_results,
                },
                f,
                indent=2,
                ensure_ascii=False,
            )
    return filepath


def export_excel(results: list, output_dir: str = None) -> str:
    """
    Export results to a multi-sheet Excel file (.xlsx) using openpyxl.
    Sheets: Master, Women, Men, Co-living, Hostel-PG-Both, Remaining.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    out_dir = output_dir or OUTPUT_DIR
    os.makedirs(out_dir, exist_ok=True)
    filepath = os.path.join(out_dir, OUTPUT_XLSX)

    wb = Workbook()

    sheets_config = [
        {"name": "Master", "filter_fn": lambda r: True},
        {"name": "Women", "filter_fn": lambda r: r.get("classification") == "women"},
        {"name": "Men", "filter_fn": lambda r: r.get("classification") == "men"},
        {"name": "Co-living", "filter_fn": lambda r: r.get("classification") == "colive"},
        {"name": "Hostel-PG-Both", "filter_fn": lambda r: r.get("classification") == "both"},
        {"name": "Remaining", "filter_fn": lambda r: r.get("classification") == "remaining"},
    ]

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2A4D69", end_color="2A4D69", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")

    data_font = Font(name="Segoe UI", size=10)

    for i, sheet_info in enumerate(sheets_config):
        sheet_name = sheet_info["name"]
        filter_fn = sheet_info["filter_fn"]

        sheet_data = [r for r in results if filter_fn(r)]

        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        ws.freeze_panes = "A2"
        ws.views.sheetView[0].showGridLines = True

        headers = [col.replace("_", " ").title() for col in CSV_COLUMNS]
        ws.append(headers)

        for col_num in range(1, len(CSV_COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        for item in sheet_data:
            row_data = [item.get(col, "") for col in CSV_COLUMNS]
            ws.append(row_data)

        max_row = ws.max_row
        for r_num in range(2, max_row + 1):
            for c_num in range(1, len(CSV_COLUMNS) + 1):
                cell = ws.cell(row=r_num, column=c_num)
                cell.font = data_font
                cell.alignment = align_left

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    wb.save(filepath)
    logger.info(f"[EXCEL] Excel exported: {filepath} ({len(results)} rows)")
    return filepath


def load_existing_results(output_dir: str = None) -> list:
    """
    Load previously scraped results. Attempts Supabase first, falls back to local JSON.
    """
    from supabase_manager import fetch_supabase_listings

    # Try cloud fetch first
    cloud_listings = fetch_supabase_listings()
    if cloud_listings:
        # Save cloud data locally for synchronization
        export_csv(cloud_listings, output_dir)
        export_json(cloud_listings, output_dir)
        export_excel(cloud_listings, output_dir)
        return cloud_listings

    # Local fallback
    out_dir = output_dir or OUTPUT_DIR
    filepath = os.path.join(out_dir, OUTPUT_JSON)

    if not os.path.exists(filepath):
        return []

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
            listings = data.get("listings", [])
            logger.info(f"[DIR] Loaded {len(listings)} existing listings from local {filepath}")
            return listings
    except Exception as e:
        logger.warning(f"[!] Could not load existing results: {e}")
        return []


def sync_local_files_with_supabase(local_results: list, output_dir: str = None) -> list:
    """
    Perform bidirectional synchronization.
    1. Fetches the latest master list from Supabase cloud database.
    2. Merges local in-memory listings with the cloud dataset, deduplicating.
    3. Re-exports updated unified CSV, JSON, and Excel files.
    4. Returns the merged master list to keep memory cache accurate.
    """
    from supabase_manager import fetch_supabase_listings

    logger.info("[SYNC] Synchronizing local files with Supabase Cloud DB...")

    # 1. Fetch latest master list from Supabase
    cloud_listings = fetch_supabase_listings()

    if not cloud_listings:
        # If Supabase is down or not configured, fallback to standard local save
        logger.warning("[SYNC] Cloud sync unavailable. Saving local cache only.")
        export_csv(local_results, output_dir)
        export_json(local_results, output_dir)
        export_excel(local_results, output_dir)
        return local_results

    # 2. Merge local cache with cloud records
    merged = merge_results(cloud_listings, local_results)

    # 3. Export all files
    export_csv(merged, output_dir)
    export_json(merged, output_dir)
    export_excel(merged, output_dir)

    logger.info(f"[SYNC] Sync complete: local files updated with {len(merged)} total global listings")
    return merged


def print_summary(results: list) -> None:
    """Print a human-readable summary of scraped results."""
    if not results:
        print("\n[STAT] No results to summarise.")
        return

    print(f"\n{'='*60}")
    print(f"[STAT] SCRAPING SUMMARY")
    print(f"{'='*60}")
    print(f"  Total listings:  {len(results)}")

    # Count by area
    areas = {}
    for r in results:
        area = r.get("area", "Unknown")
        areas[area] = areas.get(area, 0) + 1

    print(f"  Areas covered:   {len(areas)}")
    for area, cnt in sorted(areas.items(), key=lambda x: -x[1]):
        print(f"    - {area}: {cnt}")

    # Count by classification
    classifications = {}
    for r in results:
        cls = r.get("classification", "remaining")
        classifications[cls] = classifications.get(cls, 0) + 1

    print(f"\n  Classification Breakdown:")
    for cls, cnt in sorted(classifications.items(), key=lambda x: -x[1]):
        print(f"    - {cls.upper()}: {cnt}")

    # Data completeness
    with_phone = sum(1 for r in results if r.get("phone"))
    with_website = sum(1 for r in results if r.get("website"))
    with_rating = sum(1 for r in results if r.get("rating"))
    with_address = sum(1 for r in results if r.get("address_full"))
    with_coords = sum(1 for r in results if r.get("latitude") and r.get("longitude"))

    print(f"\n  Data completeness:")
    print(f"    [PHONE] Phone:       {with_phone}/{len(results)} ({100*with_phone//len(results)}%)")
    print(f"    [NET] Website:     {with_website}/{len(results)} ({100*with_website//len(results)}%)")
    print(f"    [STAR] Rating:      {with_rating}/{len(results)} ({100*with_rating//len(results)}%)")
    print(f"    [PIN] Address:     {with_address}/{len(results)} ({100*with_address//len(results)}%)")
    print(f"    [MAP] Coordinates: {with_coords}/{len(results)} ({100*with_coords//len(results)}%)")
    print(f"{'='*60}\n")
